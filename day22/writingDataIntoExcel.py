import openpyxl

#writing single data in all rows and columns
'''file="C:\\Users\\tajudeen.busari\\cameraSettings.xlsx"
workBook=openpyxl.load_workbook(file)
sheet=workBook["Sheet2"] #you can use 'sheet=workbook.active' if it is ony one sheet you r dealing with
for r in range(1,6):
    for c in range (1,4):
        sheet.cell(r,c).value="welcome"
workBook.save(file) #save the data added'''

#multiple data
file="C:\\Users\\tajudeen.busari\\cameraSettings.xlsx"
workBook=openpyxl.load_workbook(file)
sheet=workBook["Sheet3"]
sheet.cell(1,1).value=123
sheet.cell(1,2).value="taju"
sheet.cell(1,3).value="Busari"

sheet.cell(2,1).value=124
sheet.cell(2,2).value="idowu"
sheet.cell(2,3).value="Arigbabuwo"

sheet.cell(3,1).value=122
sheet.cell(3,2).value="taye"
sheet.cell(3,3).value="kehinde"

workBook.save(file)

