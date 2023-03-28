import openpyxl

#file--->workbook---->sheets--->rows--->cells
file="C:\\Users\\tajudeen.busari\\cameraSettings.xlsx"
workBook=openpyxl.load_workbook(file)
sheet=workBook["Sheet1"]
rows=sheet.max_row #no of row in the excel sheet 6 but in the range function, it will stop at 5
cols=sheet.max_column

#Reading all the rows & column from excel sheet
for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r,c).value, end="     ") #it prints in tabular form
    print()